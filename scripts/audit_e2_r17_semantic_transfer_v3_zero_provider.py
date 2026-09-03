#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import inspect
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from research_pipeline.e2_r17_controlled_suite_schema import answer_cells, sha256_file
from research_pipeline.e2_r17_semantic_transfer_v3_builders import (
    BINDING,
    EXPECTED_GENERATION_RUNTIME,
    PROCEDURAL,
    SEMANTIC_TYPES,
    SKELETONS,
    generation_runtime_fingerprint,
    observable_route,
    visible_router_features,
)

EXPECTED = {
    "suite_manifest": "9d57c0abc51758e3657484048e9a132a531ff2758d724b7be5cc6d14ae262338",
    "split": "815977e908214b66a1106d623ca68f4707d56b117fb01740cadbd1edeab3679e",
    "metadata": "1cd8fbe40ab84d9db32a6b4877a6aeb3949b4db0772cba04bf9d60ca901b612f",
    "dataset": "be84cca6d75359b713a1d6f914c002f7f2be95bcef5f4b745e61908ac7d56b10",
}

FORBIDDEN_PROMPT_MARKERS = (
    "PROCEDURAL_TRANSFORMATION",
    "INSTANCE_BINDING_LOCALIZATION",
    "cross_join_ledger",
    "cross_measure_panel",
    "cross_snapshot_bundle",
    "cross_group_window",
    "cross_lookup_reconcile",
    "semantic_type",
    "matched_skeleton",
    "reusable_transform_steps",
    "binding_candidate_count",
)


def req(condition: bool, message: str) -> None:
    if not condition:
        raise RuntimeError(message)


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def workbook_diff_only_answers(init_path: Path, golden_path: Path, answer_position: str, expected: dict[str, Any]) -> None:
    init = load_workbook(init_path, data_only=False)
    golden = load_workbook(golden_path, data_only=False)
    try:
        req(init.sheetnames == golden.sheetnames, f"sheet drift: {init_path}")
        allowed = {(sheet, cell) for sheet, cell in answer_cells(answer_position)}
        for sheet in init.sheetnames:
            iw = init[sheet]
            gw = golden[sheet]
            for row in range(1, max(iw.max_row, gw.max_row) + 1):
                for col in range(1, max(iw.max_column, gw.max_column) + 1):
                    ic = iw.cell(row=row, column=col)
                    gc = gw.cell(row=row, column=col)
                    key = (sheet, ic.coordinate)
                    if key in allowed:
                        req(ic.value is None, f"init answer not blank: {init_path} {key}")
                        qualified = f"{sheet}!{ic.coordinate}"
                        req(qualified in expected, f"missing expected answer: {qualified}")
                        exp = expected[qualified]
                        if isinstance(gc.value, (int, float)) and isinstance(exp, (int, float)):
                            req(abs(float(gc.value) - float(exp)) <= 1e-9, f"golden numeric answer drift: {golden_path} {qualified}")
                        else:
                            req(gc.value == exp, f"golden answer drift: {golden_path} {qualified}")
                    else:
                        req(ic.value == gc.value, f"non-answer cell differs: {init_path} {sheet}!{ic.coordinate}")
    finally:
        init.close()
        golden.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--suite-root", type=Path, required=True)
    parser.add_argument("--rebuild-root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    req(not args.output.exists(), "audit output already exists")

    paths = {
        "suite_manifest": args.suite_root / "suite_manifest.json",
        "split": args.suite_root / "r17_semantic_transfer_v3_split_manifest.json",
        "metadata": args.suite_root / "r17_semantic_transfer_v3_metadata.json",
        "dataset": args.suite_root / "spreadsheetbench_verified_400" / "dataset.json",
    }
    rebuild_paths = {
        "suite_manifest": args.rebuild_root / "suite_manifest.json",
        "split": args.rebuild_root / "r17_semantic_transfer_v3_split_manifest.json",
        "metadata": args.rebuild_root / "r17_semantic_transfer_v3_metadata.json",
        "dataset": args.rebuild_root / "spreadsheetbench_verified_400" / "dataset.json",
    }
    for key, path in paths.items():
        req(path.is_file(), f"missing {key}: {path}")
        req(sha(path) == EXPECTED[key], f"frozen {key} hash drift")
        req(rebuild_paths[key].is_file(), f"missing deterministic rebuild {key}")
        req(sha(rebuild_paths[key]) == sha(path), f"deterministic rebuild mismatch: {key}")

    manifest = load(paths["suite_manifest"])
    audit_runtime = generation_runtime_fingerprint()
    req(audit_runtime == EXPECTED_GENERATION_RUNTIME, f"audit generation runtime drift: {audit_runtime}")
    req(manifest.get("generation_runtime_pinned") is True, "suite generation runtime is not pinned")
    req(manifest.get("generation_runtime") == EXPECTED_GENERATION_RUNTIME, "suite generation runtime fingerprint drift")
    split = load(paths["split"])
    metadata_rows = load(paths["metadata"])
    dataset_rows = load(paths["dataset"])
    metadata = {str(row["id"]): row for row in metadata_rows}
    dataset = {str(row["id"]): row for row in dataset_rows}

    req(manifest["status"] == "PASS_ZERO_PROVIDER_SEMANTIC_TRANSFER_V3_CROSSED_SUITE_MATERIALIZATION", "suite status drift")
    req(manifest["provider_calls"] == 0 and manifest["scientific_outcomes_accessed"] is False, "provider/outcome boundary crossed")
    req(manifest["task_count"] == 270, "generated task count drift")
    req(manifest["update_streams"] == 20 and manifest["update_tasks"] == 160, "update shape drift")
    req(manifest["heldout_tasks"] == 20, "heldout shape drift")
    req(manifest["crossed_cells"] == 10 and manifest["streams_per_crossed_cell"] == 2, "crossed-cell shape drift")
    req(manifest["pair_init_checks"] == 135 and manifest["pair_init_mismatches"] == 0, "crossed init invariant drift")
    req(manifest["observable_route_hidden_label_mismatches"] == 0, "router/hidden-label qualification mismatch")
    req(manifest["observable_route_counts_all_generated_tasks"] == {"MRW4": 135, "UNCLASSIFIED": 0, "WIN-C": 135}, "router balance drift")
    req(all(v == 0 for v in manifest["old_suite_disjointness"]["task_id_overlap"].values()), "old task ID overlap")
    req(all(v == 0 for v in manifest["old_suite_disjointness"]["xlsx_sha256_overlap"].values()), "old XLSX overlap")

    # Router code contract: instruction is the only argument. No metadata object
    # can be passed even accidentally without changing this audited signature.
    req(list(inspect.signature(observable_route).parameters) == ["instruction"], "observable route gained privileged input")
    req(list(inspect.signature(visible_router_features).parameters) == ["instruction"], "visible features gained privileged input")

    prompt_leaks: list[dict[str, str]] = []
    router_mismatches: list[str] = []
    instruction_lengths: dict[str, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))
    for task_id, record in dataset.items():
        instruction = str(record["instruction"])
        lower = instruction.lower()
        for marker in FORBIDDEN_PROMPT_MARKERS:
            if marker.lower() in lower:
                prompt_leaks.append({"task_id": task_id, "marker": marker})
        row = metadata[task_id]
        route = observable_route(instruction)
        expected = "MRW4" if row["semantic_type"] == PROCEDURAL else "WIN-C"
        if route != expected or route != row["observable_router_route"]:
            router_mismatches.append(task_id)
        instruction_lengths[row["matched_skeleton"]][row["semantic_type"]].append(len(instruction.split()))
    req(not prompt_leaks, f"model-visible hidden label leak: {prompt_leaks[:5]}")
    req(not router_mismatches, f"router mismatch: {router_mismatches[:5]}")

    # The intended semantic manipulation necessarily changes operation/binding
    # wording, but paired prompt lengths must remain narrow rather than creating
    # a gross information-budget confound.
    length_balance: dict[str, dict[str, float]] = {}
    for skeleton in SKELETONS:
        proc = instruction_lengths[skeleton][PROCEDURAL]
        bind = instruction_lengths[skeleton][BINDING]
        proc_mean = sum(proc) / len(proc)
        bind_mean = sum(bind) / len(bind)
        diff = abs(proc_mean - bind_mean)
        req(diff <= 8.0, f"instruction length imbalance too large: {skeleton} diff={diff}")
        length_balance[skeleton] = {"procedural_mean_words": proc_mean, "binding_mean_words": bind_mean, "absolute_difference_words": diff}

    streams = {str(k): list(map(str, v)) for k, v in split["update_streams"].items()}
    req(len(streams) == 20 and all(len(v) == 8 for v in streams.values()), "stream cardinality drift")
    cell_streams: dict[tuple[str, str], list[str]] = defaultdict(list)
    profiles_by_cell_block: dict[tuple[str, str, int], tuple[int, ...]] = {}
    update_ids: list[str] = []
    for sid, tids in streams.items():
        update_ids.extend(tids)
        rows = [metadata[x] for x in tids]
        skeletons = {str(r["matched_skeleton"]) for r in rows}
        semantics = {str(r["semantic_type"]) for r in rows}
        blocks = {int(r["block"]) for r in rows}
        req(len(skeletons) == len(semantics) == len(blocks) == 1, f"stream homogeneity drift: {sid}")
        skeleton = next(iter(skeletons)); semantic = next(iter(semantics)); block = next(iter(blocks))
        cell_streams[(skeleton, semantic)].append(sid)
        profiles_by_cell_block[(skeleton, semantic, block)] = tuple(sorted(int(r["profile_index"]) for r in rows))
    req(len(update_ids) == 160 and len(set(update_ids)) == 160, "update task uniqueness drift")
    req(len(cell_streams) == 10 and all(len(v) == 2 for v in cell_streams.values()), "cell stream balance drift")

    # Strong selection crossing: the scientific profile set is identical across
    # semantic cells within every skeleton and update block.
    for skeleton in SKELETONS:
        for block in (21, 22):
            req(
                profiles_by_cell_block[(skeleton, PROCEDURAL, block)] == profiles_by_cell_block[(skeleton, BINDING, block)],
                f"semantic-blind update profile mismatch: {skeleton} b{block}",
            )

    heldout = list(map(str, split["common_heldout_probe"]))
    req(len(heldout) == 20 and len(set(heldout)) == 20 and set(heldout).isdisjoint(update_ids), "heldout shape/overlap drift")
    heldout_profiles: dict[tuple[str, str], tuple[int, ...]] = {}
    for skeleton in SKELETONS:
        for semantic in SEMANTIC_TYPES:
            rows = [metadata[x] for x in heldout if metadata[x]["matched_skeleton"] == skeleton and metadata[x]["semantic_type"] == semantic]
            req(len(rows) == 2 and all(int(r["block"]) == 23 for r in rows), f"heldout cell drift: {skeleton} {semantic}")
            heldout_profiles[(skeleton, semantic)] = tuple(sorted(int(r["profile_index"]) for r in rows))
        req(heldout_profiles[(skeleton, PROCEDURAL)] == heldout_profiles[(skeleton, BINDING)], f"semantic-blind heldout profile mismatch: {skeleton}")

    verified = args.suite_root / "spreadsheetbench_verified_400" / "spreadsheet"
    pair_init_hash: dict[str, dict[str, str]] = defaultdict(dict)
    workbook_pairs_checked = 0
    for task_id, row in metadata.items():
        task_dir = verified / task_id
        init_path = task_dir / f"{task_id}_init.xlsx"
        golden_path = task_dir / f"{task_id}_golden.xlsx"
        req(init_path.is_file() and golden_path.is_file(), f"missing workbook: {task_id}")
        workbook_diff_only_answers(init_path, golden_path, dataset[task_id]["answer_position"], row["golden_answer_cells"])
        pair_init_hash[str(row["pair_key"])][str(row["semantic_type"])] = sha256_file(init_path)
        workbook_pairs_checked += 1
    req(workbook_pairs_checked == 270, "workbook validation count drift")
    req(len(pair_init_hash) == 135, "pair-key cardinality drift")
    for key, values in pair_init_hash.items():
        req(set(values) == set(SEMANTIC_TYPES), f"pair missing semantic side: {key}")
        req(len(set(values.values())) == 1, f"paired init bytes differ: {key}")

    # Exact MRW4 treatment text must be frozen in the split, rather than
    # delegated to an unspecified historical rule.
    req(split["mrw4_failed_witness_selector"] == "lowest original rollout index among verifier-failure nonwinner trajectories", "MRW witness selector underspecified")
    req("semantic-transfer-mrw4-v3" in split["treated_pool_selection"], "MRW treated-pool salt drift")
    req(split["hidden_semantic_labels_are_router_inputs"] is False, "hidden labels admitted to router")
    req(split["scientific_unit"] == "matched_skeleton_interaction", "scientific unit drift")

    payload = {
        "schema_version": "1.0",
        "artifact_type": "e2-r17-semantic-transfer-v3-zero-provider-audit",
        "status": "PASS_V3_CROSSED_IDENTIFICATION_ZERO_PROVIDER_AUDIT",
        "provider_calls": 0,
        "scientific_execution": False,
        "scientific_outcomes_accessed": False,
        "generation_runtime": {
            "expected": EXPECTED_GENERATION_RUNTIME,
            "suite": manifest["generation_runtime"],
            "audit": audit_runtime,
            "match": True,
        },
        "frozen_hashes": {key: sha(path) for key, path in paths.items()},
        "deterministic_rebuild_match": True,
        "suite": {
            "generated_tasks": 270,
            "update_streams": 20,
            "update_tasks": 160,
            "heldout_tasks": 20,
            "crossed_skeletons": 5,
            "crossed_cells": 10,
            "paired_init_checks": 135,
            "paired_init_mismatches": 0,
            "workbook_pairs_checked": workbook_pairs_checked,
            "old_task_id_overlap": 0,
            "old_xlsx_sha256_overlap": 0,
        },
        "identification": {
            "semantic_crossed_inside_common_generator": True,
            "scientific_profile_selection_semantic_blind": True,
            "heldout_profile_selection_semantic_blind": True,
            "primary_independent_units": 5,
            "primary_unit": "matched_skeleton_interaction I_h",
            "minimum_all_positive_exact_sign_p": 0.03125,
            "replicates_are_measurement_not_semantic_replication": True,
            "within_stream_sd_power_claim_removed": True,
        },
        "router": {
            "input": "actor-visible instruction only",
            "forbidden_hidden_metadata_inputs": True,
            "generated_task_routes": {"MRW4": 135, "WIN-C": 135, "UNCLASSIFIED": 0},
            "hidden_label_mismatches": 0,
            "prompt_hidden_label_leaks": 0,
            "instruction_length_balance": length_balance,
        },
        "treatment": {
            "failed_witness_selector": split["mrw4_failed_witness_selector"],
            "treated_pool_selection": split["treated_pool_selection"],
            "exactly_four_treated_pools_per_support_qualified_stream": True,
        },
        "authority": {
            "stage_a_provider_execution": False,
            "stage_b_learning_execution": False,
            "heldout_evaluation": False,
            "analyzer": False,
            "paper_promotion": False,
            "second_backbone": False,
            "public_benchmark": False,
        },
        "next_gate": "V3_CONTROL_PLANE_IMPLEMENTATION_AND_INDEPENDENT_PRE_STAGE_A_REVIEW",
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(payload, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
