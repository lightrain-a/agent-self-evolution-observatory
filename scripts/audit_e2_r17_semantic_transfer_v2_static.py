#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from research_pipeline.e2_r17_controlled_suite_schema import answer_cells, sha256_file
from research_pipeline.e2_r17_semantic_transfer_builders import FAMILIES, FAMILY_SPECS

OLD_FAMILIES = {
    "aggregation_join",
    "formula_materialization",
    "input_output_contract",
    "multi_step_pipeline",
    "schema_key_alignment",
    "target_sheet_range",
}
EXPECTED_SUITE_SHA = "6ceabf7607856214984fff2c39cdf8ceb8a02620c519d9790a89ecefefe91071"
EXPECTED_SPLIT_SHA = "f6c3a5e24cc284e27006d9a9fbc2f0d20f185ec53e0e6f58d05a7bd5a0afdd18"
EXPECTED_METADATA_SHA = "5fb76d7da7ff19e67fdaad8e3a8c1c4dbb82337e524d31b450cdf7f91c62d5e8"


def req(condition: bool, message: str) -> None:
    if not condition:
        raise RuntimeError(message)


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def workbook_diff_only_answers(init_path: Path, golden_path: Path, answer_position: str, expected: dict[str, Any]) -> None:
    init = load_workbook(init_path, data_only=False)
    golden = load_workbook(golden_path, data_only=False)
    try:
        req(init.sheetnames == golden.sheetnames, f"sheet drift: {init_path}")
        allowed = {(sheet, cell) for sheet, cell in answer_cells(answer_position)}
        for sheet in init.sheetnames:
            iw = init[sheet]
            gw = golden[sheet]
            max_row = max(iw.max_row, gw.max_row)
            max_col = max(iw.max_column, gw.max_column)
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = iw.cell(row=row, column=col)
                    gcell = gw.cell(row=row, column=col)
                    key = (sheet, cell.coordinate)
                    if key in allowed:
                        req(cell.value is None, f"init answer not blank: {init_path} {key}")
                        qualified = f"{sheet}!{cell.coordinate}"
                        req(qualified in expected, f"missing expected answer: {qualified}")
                        req(gcell.value == expected[qualified], f"golden answer drift: {golden_path} {qualified}")
                    else:
                        req(cell.value == gcell.value, f"non-answer cell differs: {init_path} {sheet}!{cell.coordinate}")
    finally:
        init.close()
        golden.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--suite-root", type=Path, required=True)
    parser.add_argument("--e2-root", type=Path, required=True)
    parser.add_argument("--prior-v1-audit", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    req(not args.output.exists(), "static audit output already exists")
    suite_manifest_path = args.suite_root / "suite_manifest.json"
    split_path = args.suite_root / "r17_semantic_transfer_split_manifest.json"
    metadata_path = args.suite_root / "r17_semantic_transfer_metadata.json"
    dataset_path = args.suite_root / "spreadsheetbench_verified_400" / "dataset.json"
    for path in (suite_manifest_path, split_path, metadata_path, dataset_path, args.prior_v1_audit):
        req(path.is_file(), f"missing artifact: {path}")

    req(sha(suite_manifest_path) == EXPECTED_SUITE_SHA, "semantic-transfer suite manifest drift")
    req(sha(split_path) == EXPECTED_SPLIT_SHA, "semantic-transfer split drift")
    req(sha(metadata_path) == EXPECTED_METADATA_SHA, "semantic-transfer metadata drift")

    manifest = load_json(suite_manifest_path)
    split = load_json(split_path)
    metadata_rows = load_json(metadata_path)
    dataset_rows = load_json(dataset_path)
    prior_v1 = load_json(args.prior_v1_audit)
    metadata = {str(row["id"]): row for row in metadata_rows}
    dataset = {str(row["id"]): row for row in dataset_rows}

    req(manifest["status"] == "PASS_ZERO_PROVIDER_SEMANTIC_TRANSFER_SUITE_MATERIALIZATION", "suite status drift")
    req(manifest.get("provider_calls") == 0 and manifest.get("scientific_outcomes_accessed") is False, "suite crossed provider/outcome boundary")
    req(int(manifest["task_count"]) == 216, "task count drift")
    req(int(manifest["update_streams"]) == 18 and int(manifest["update_tasks"]) == 144, "update shape drift")
    req(int(manifest["heldout_tasks"]) == 18, "heldout shape drift")
    req(set(FAMILIES).isdisjoint(OLD_FAMILIES), "new family identity collides with closed family")
    req(set(manifest["families"]) == set(FAMILIES), "family universe drift")
    req(all(v == 0 for v in manifest["old_suite_disjointness"]["task_id_overlap"].values()), "old task ID overlap")
    req(all(v == 0 for v in manifest["old_suite_disjointness"]["xlsx_sha256_overlap"].values()), "old XLSX overlap")

    route = split["semantic_routing_rule"]
    req(route["PROCEDURAL_TRANSFORMATION"] == "MRW4", "procedural route drift")
    req(route["INSTANCE_BINDING_LOCALIZATION"] == "WIN-C", "binding route drift")

    family_semantics: dict[str, str] = {}
    family_skeletons: dict[str, str] = {}
    for family, spec in FAMILY_SPECS.items():
        family_semantics[family] = str(spec["semantic_type"])
        family_skeletons[family] = str(spec["matched_skeleton"])
        if spec["semantic_type"] == "PROCEDURAL_TRANSFORMATION":
            req(int(spec["reusable_transform_steps"]) >= 2 and int(spec["binding_candidate_count"]) == 1, f"procedural mechanical rule drift: {family}")
        else:
            req(int(spec["binding_candidate_count"]) >= 2 and int(spec["reusable_transform_steps"]) <= 1, f"binding mechanical rule drift: {family}")

    skeleton_to_semantics: dict[str, set[str]] = {}
    for family in FAMILIES:
        skeleton_to_semantics.setdefault(family_skeletons[family], set()).add(family_semantics[family])
    req(len(skeleton_to_semantics) == 3, "matched skeleton count drift")
    req(all(values == {"PROCEDURAL_TRANSFORMATION", "INSTANCE_BINDING_LOCALIZATION"} for values in skeleton_to_semantics.values()), "matched skeleton semantic crossing drift")

    streams = {str(k): list(map(str, v)) for k, v in split["update_streams"].items()}
    req(len(streams) == 18 and all(len(tasks) == 8 for tasks in streams.values()), "stream cardinality drift")
    update_ids = [task for tasks in streams.values() for task in tasks]
    req(len(update_ids) == 144 and len(set(update_ids)) == 144, "update task uniqueness drift")
    stream_semantic_counts = {"PROCEDURAL_TRANSFORMATION": 0, "INSTANCE_BINDING_LOCALIZATION": 0}
    stream_skeleton_counts: dict[str, int] = {}
    for stream_id, task_ids in streams.items():
        rows = [metadata[task_id] for task_id in task_ids]
        families = {row["primary_failure_family"] for row in rows}
        semantics = {row["semantic_type"] for row in rows}
        skeletons = {row["matched_skeleton"] for row in rows}
        req(len(families) == len(semantics) == len(skeletons) == 1, f"stream not homogeneous: {stream_id}")
        semantic = str(next(iter(semantics)))
        skeleton = str(next(iter(skeletons)))
        stream_semantic_counts[semantic] += 1
        stream_skeleton_counts[skeleton] = stream_skeleton_counts.get(skeleton, 0) + 1
    req(stream_semantic_counts == {"PROCEDURAL_TRANSFORMATION": 9, "INSTANCE_BINDING_LOCALIZATION": 9}, "semantic stream balance drift")
    req(all(value == 6 for value in stream_skeleton_counts.values()) and len(stream_skeleton_counts) == 3, "skeleton stream balance drift")

    heldout = list(map(str, split["common_heldout_probe"]))
    req(len(heldout) == 18 and len(set(heldout)) == 18, "heldout uniqueness drift")
    req(set(heldout).isdisjoint(update_ids), "heldout/update overlap")
    heldout_family_counts = {family: 0 for family in FAMILIES}
    heldout_semantic_counts = {"PROCEDURAL_TRANSFORMATION": 0, "INSTANCE_BINDING_LOCALIZATION": 0}
    for task_id in heldout:
        row = metadata[task_id]
        req(int(row["block"]) == 20, f"heldout block drift: {task_id}")
        heldout_family_counts[str(row["primary_failure_family"])] += 1
        heldout_semantic_counts[str(row["semantic_type"])] += 1
    req(all(value == 3 for value in heldout_family_counts.values()), "heldout family balance drift")
    req(heldout_semantic_counts == {"PROCEDURAL_TRANSFORMATION": 9, "INSTANCE_BINDING_LOCALIZATION": 9}, "heldout semantic balance drift")

    verified_root = args.suite_root / "spreadsheetbench_verified_400" / "spreadsheet"
    workbook_pairs_checked = 0
    for task_id, row in metadata.items():
        record = dataset[task_id]
        task_dir = verified_root / task_id
        init_path = task_dir / f"{task_id}_init.xlsx"
        golden_path = task_dir / f"{task_id}_golden.xlsx"
        req(init_path.is_file() and golden_path.is_file(), f"missing workbook pair: {task_id}")
        workbook_diff_only_answers(init_path, golden_path, str(record["answer_position"]), row["golden_answer_cells"])
        workbook_pairs_checked += 1
    req(workbook_pairs_checked == 216, "workbook validation count drift")

    # Prove this new task namespace is untouched by prior scientific run artifacts.
    code_re = "|".join(re.escape(str(FAMILY_SPECS[family]["code"])) for family in FAMILIES)
    task_pattern = rf"r17-b(17|18|19|20)-({code_re})-p[0-8]"
    runs_root = args.e2_root / "runs"
    rg = subprocess.run(["rg", "-l", task_pattern, str(runs_root)], capture_output=True, text=True)
    req(rg.returncode in (0, 1), "run artifact scan failed")
    historical_run_refs = [line for line in rg.stdout.splitlines() if line.strip()] if rg.returncode == 0 else []
    req(not historical_run_refs, f"new semantic-transfer task found in historical run artifact: {len(historical_run_refs)}")
    trajectory_refs = []
    for path in args.e2_root.rglob("r17_trajectory_ref.json"):
        text = path.read_text(encoding="utf-8", errors="ignore")
        if re.search(task_pattern, text):
            trajectory_refs.append(str(path))
    req(not trajectory_refs, f"new semantic-transfer trajectory refs exist: {len(trajectory_refs)}")

    req(prior_v1["status"] == "PASS_SEMANTIC_TRANSFER_V1_ZERO_PROVIDER_STATIC_AUDIT", "prior semantic-transfer V1 static status drift")
    req(prior_v1.get("provider_calls") == 0 and prior_v1.get("scientific_execution") is False, "prior V1 was not zero-provider")
    req(prior_v1.get("new_test_outcomes_accessed") is False, "prior V1 outcome boundary crossed")

    payload = {
        "schema_version": "1.0",
        "artifact_type": "e2-r17-selective-mrw-semantic-transfer-v2-static-audit",
        "status": "PASS_SEMANTIC_TRANSFER_V2_ZERO_PROVIDER_STATIC_AUDIT",
        "provider_calls": 0,
        "scientific_execution": False,
        "new_test_outcomes_accessed": False,
        "suite": {
            "root": str(args.suite_root),
            "suite_manifest_sha256": sha(suite_manifest_path),
            "split_manifest_sha256": sha(split_path),
            "metadata_sha256": sha(metadata_path),
            "dataset_sha256": sha(dataset_path),
            "task_count": 216,
            "update_streams": 18,
            "update_tasks": 144,
            "heldout_tasks": 18,
            "workbook_pairs_checked": workbook_pairs_checked,
            "old_task_id_overlap": 0,
            "old_xlsx_sha256_overlap": 0,
            "historical_run_refs": 0,
            "historical_trajectory_refs": 0,
        },
        "semantic_identification": {
            "old_family_ids_disjoint": True,
            "mechanical_rule_bound": True,
            "procedural_streams": 9,
            "binding_streams": 9,
            "matched_skeletons": {key: sorted(value) for key, value in skeleton_to_semantics.items()},
            "family_semantics": family_semantics,
            "family_skeletons": family_skeletons,
            "family_id_lookup_from_closed_sample_can_route_test": False,
            "equal_dose_stage_b_requires_four_mixed_pools_per_stream": True,
        },
        "prior_semantic_transfer_v1": {
            "path": str(args.prior_v1_audit),
            "sha256": sha(args.prior_v1_audit),
            "provider_calls": 0,
            "new_test_outcomes_accessed": False,
            "disposition": "SUPERSEDED_PRE_PROVIDER_BY_SEMANTIC_TRANSFER_V2_POWER_ALLOCATION_REPAIR",
        },
        "authority": {
            "current_provider_identity_qualification": True,
            "zero_provider_stage_a_preflight": True,
            "stage_a_provider_execution": False,
            "stage_b_learning_execution": False,
            "heldout_evaluation": False,
            "analyzer": False,
            "second_backbone": False,
            "public_benchmark": False,
            "paper_promotion": False,
        },
        "next_gate": "CURRENT_DEEPSEEK_IDENTITY_QUALIFICATION_THEN_ZERO_PROVIDER_STAGE_A_PREFLIGHT",
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({"status": payload["status"], "suite": payload["suite"], "semantic_identification": payload["semantic_identification"], "next_gate": payload["next_gate"]}, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
