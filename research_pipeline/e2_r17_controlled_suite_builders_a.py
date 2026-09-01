from __future__ import annotations

import random
from typing import Any

from openpyxl import Workbook

from .e2_r17_controlled_suite_schema import answer_range


def build_input_output_contract(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    ws = wb.create_sheet("ContractSource")
    headers = ["record_id", "amount", "status"]
    if ambiguity >= 1:
        headers.append("amount_backup")
    if ambiguity >= 2:
        headers.extend(["status_previous", "record_id_old"])
    ws.append(headers)
    values: list[int] = []
    for index in range(1, 8 + depth * 2):
        amount = rng.randint(5, 90)
        values.append(amount)
        row: list[Any] = [f"R{index:02d}", amount, "active" if index % 3 else "inactive"]
        if ambiguity >= 1:
            row.append(amount + 1000)
        if ambiguity >= 2:
            row.extend(["active", f"OLD{index:02d}"])
        ws.append(row)
    sentinel = f"KEEP-{task_id}"
    ws["F1"] = sentinel
    checksum = sum(values)
    active_sum = sum(value for index, value in enumerate(values, start=1) if index % 3)
    result = wb["Result"]
    result["B2"] = checksum
    result["B3"] = sentinel
    last = 3
    if depth >= 1:
        result["B4"] = active_sum
        last = 4
    if depth >= 2:
        result["B5"] = len(values)
        last = 5
    instruction = (
        "Start from an exact copy of input.xlsx and preserve every existing worksheet and cell. "
        f"In Result!B2 write the materialized sum of ContractSource!B2:B{len(values)+1}; "
        "in Result!B3 copy the immutable sentinel from ContractSource!F1. "
    )
    if depth >= 1:
        instruction += "In Result!B4 write the sum of Amount for rows whose Status is active. "
    if depth >= 2:
        instruction += "In Result!B5 write the number of data rows. "
    instruction += "Save only as output.xlsx; do not alter input.xlsx and do not leave formulas."
    return instruction, answer_range(last), {
        "checksum": checksum,
        "sentinel": sentinel,
        "active_sum": active_sum,
        "row_count": len(values),
    }


def build_target_sheet_range(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    target_name = f"Quarter_{(int(task_id[-1], 36) % 4) + 1}_FINAL"
    candidates = [target_name]
    if ambiguity >= 1:
        candidates.append(target_name.replace("FINAL", "DRAFT"))
    if ambiguity >= 2:
        candidates.extend([target_name.replace("FINAL", "ARCHIVE"), target_name + "_COPY"])
    target_values: list[int] = []
    for sheet_index, name in enumerate(candidates):
        ws = wb.create_sheet(name)
        ws.append(["item", "net_value", "net_value_old" if ambiguity else "note"])
        values = [rng.randint(20, 180) + sheet_index * 500 for _ in range(6 + depth * 2)]
        if sheet_index == 0:
            target_values = values
        for index, value in enumerate(values, start=1):
            ws.append([f"I{index:02d}", value, value + 900 if ambiguity else f"row-{index}"])
    result = wb["Result"]
    result["B2"] = sum(target_values)
    result["B3"] = max(target_values)
    last = 3
    if depth >= 1:
        result["B4"] = round(sum(target_values) / len(target_values), 2)
        last = 4
    if depth >= 2:
        result["B5"] = sum(1 for value in target_values if value >= 100)
        last = 5
    instruction = (
        f"Use only worksheet '{target_name}' and ignore similarly named sheets. For net_value cells "
        f"B2:B{len(target_values)+1}, write the materialized sum to Result!B2 and maximum to Result!B3. "
    )
    if depth >= 1:
        instruction += "Write the arithmetic mean rounded to 2 decimals to Result!B4. "
    if depth >= 2:
        instruction += "Write the count of values at least 100 to Result!B5. "
    instruction += "Save as output.xlsx with values, not formulas."
    return instruction, answer_range(last), {
        "target_sheet": target_name,
        "target_values": target_values,
    }


def build_schema_key_alignment(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    del task_id
    tx = wb.create_sheet("Transactions")
    rate = wb.create_sheet("RateTable")
    key_header = ("account_code", "acct_code", "account_identifier")[ambiguity]
    amount_header = ("amount_usd", "posted_amount", "recognized_amount")[ambiguity]
    tx_headers = [key_header, amount_header, "status"]
    if ambiguity >= 1:
        tx_headers.append("amount_estimate")
    if ambiguity >= 2:
        tx_headers.append("account_identifier_old")
    tx.append(tx_headers)
    rate_key_header = ("account_code", "ledger_code", "rate_key")[ambiguity]
    rate.append([rate_key_header, "conversion_rate", "rate_archived" if ambiguity else "note"])
    codes = [f"A{index:02d}" for index in range(1, 7)]
    rates = {code: round(0.75 + index * 0.11, 2) for index, code in enumerate(codes)}
    for code in codes:
        rate.append([code, rates[code], rates[code] + 2 if ambiguity else "current"])
    converted: list[float] = []
    active_converted: list[float] = []
    for index in range(10 + depth * 3):
        code = codes[index % len(codes)]
        amount = rng.randint(20, 250)
        status = "active" if index % 4 else "inactive"
        row: list[Any] = [code, amount, status]
        if ambiguity >= 1:
            row.append(amount + 500)
        if ambiguity >= 2:
            row.append(f"OLD-{code}")
        tx.append(row)
        value = round(amount * rates[code], 2)
        converted.append(value)
        if status == "active":
            active_converted.append(value)
    result = wb["Result"]
    result["B2"] = round(sum(converted), 2)
    last = 2
    if depth >= 1:
        result["B3"] = round(sum(active_converted), 2)
        last = 3
    if depth >= 2:
        result["B4"] = max(converted)
        last = 4
    instruction = (
        f"Join Transactions.{key_header} to RateTable.{rate_key_header}. Multiply Transactions.{amount_header} "
        "by RateTable.conversion_rate and write the materialized grand total, rounded to 2 decimals, to Result!B2. "
    )
    if depth >= 1:
        instruction += "Write the corresponding total for rows whose Transactions.status is active to Result!B3. "
    if depth >= 2:
        instruction += "Write the largest converted row amount to Result!B4. "
    instruction += "Ignore old, estimate, and archive columns and save as output.xlsx."
    return instruction, answer_range(last), {
        "key_header": key_header,
        "amount_header": amount_header,
        "rate_key_header": rate_key_header,
        "converted": converted,
    }
