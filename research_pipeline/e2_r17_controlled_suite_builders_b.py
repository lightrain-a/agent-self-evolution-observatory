from __future__ import annotations

import random
from typing import Any

from openpyxl import Workbook

from .e2_r17_controlled_suite_schema import answer_range


def build_aggregation_join(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    del task_id
    orders = wb.create_sheet("Orders")
    customers = wb.create_sheet("Customers")
    products = wb.create_sheet("Products")
    customer_key = ("customer_id", "cust_id", "customer_key")[ambiguity]
    product_key = ("product_id", "sku", "product_key")[ambiguity]
    orders.append(
        ["order_id", customer_key, product_key, "quantity", "unit_price", "quantity_old" if ambiguity else "note"]
    )
    customers.append([customer_key, "region", "region_previous" if ambiguity else "note"])
    products.append([product_key, "category", "category_old" if ambiguity else "note"])
    customer_rows = [(f"C{i:02d}", "East" if i % 2 else "West") for i in range(1, 9)]
    product_rows = [(f"P{i:02d}", "Core" if i % 3 else "Other") for i in range(1, 7)]
    for key, region in customer_rows:
        customers.append([key, region, "North" if ambiguity else "current"])
    for key, category in product_rows:
        products.append([key, category, "Legacy" if ambiguity else "current"])
    regions = {key: region for key, region in customer_rows}
    categories = {key: category for key, category in product_rows}
    selected: list[float] = []
    for index in range(14 + depth * 4):
        ckey = customer_rows[index % len(customer_rows)][0]
        pkey = product_rows[(index * 2 + 1) % len(product_rows)][0]
        quantity = 1 + index % 5
        price = rng.randint(8, 80)
        orders.append([f"O{index+1:03d}", ckey, pkey, quantity, price, quantity + 100 if ambiguity else "live"])
        keep = regions[ckey] == "East"
        if depth >= 1:
            keep = keep and categories[pkey] == "Core"
        if keep:
            selected.append(float(quantity * price))
    result = wb["Result"]
    result["B2"] = round(sum(selected), 2)
    result["B3"] = len(selected)
    last = 3
    if depth >= 2:
        result["B4"] = round(sum(selected) / len(selected), 2) if selected else 0.0
        last = 4
    instruction = f"Join Orders.{customer_key} to Customers.{customer_key} and keep only region East. "
    if depth >= 1:
        instruction += f"Also join Orders.{product_key} to Products.{product_key} and keep only category Core. "
    instruction += (
        "Compute quantity * unit_price for retained orders; write the materialized revenue sum to Result!B2 "
        "and retained order count to Result!B3. "
    )
    if depth >= 2:
        instruction += "Write average retained order value rounded to 2 decimals to Result!B4. "
    instruction += "Ignore old columns and save as output.xlsx."
    return instruction, answer_range(last), {
        "retained_count": len(selected),
        "retained_values": selected,
    }


def build_formula_materialization(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    del task_id
    items = wb.create_sheet("Items")
    qty_header = ("quantity", "qty", "units_sold")[ambiguity]
    price_header = ("unit_price", "price", "recognized_unit_price")[ambiguity]
    headers = ["item_id", qty_header, price_header, "discount_rate", "tax_rate"]
    if ambiguity >= 1:
        headers.append("price_estimate")
    if ambiguity >= 2:
        headers.append("quantity_previous")
    items.append(headers)
    outputs: list[float] = []
    n_rows = 6 + depth * 2
    for index in range(n_rows):
        quantity = 1 + index % 6
        price = rng.randint(7, 65)
        discount = (index % 3) * 0.05
        tax = 0.08 + (index % 2) * 0.02
        row: list[Any] = [f"I{index+1:02d}", quantity, price, discount, tax]
        if ambiguity >= 1:
            row.append(price + 100)
        if ambiguity >= 2:
            row.append(quantity + 20)
        items.append(row)
        value = quantity * price
        if depth >= 1:
            value *= 1 - discount
        if depth >= 2:
            value *= 1 + tax
        outputs.append(round(value, 2))
    result = wb["Result"]
    for index, value in enumerate(outputs, start=2):
        result.cell(row=index, column=2, value=value)
    instruction = (
        f"For each row in Items, compute {qty_header} * {price_header}"
        + (" * (1 - discount_rate)" if depth >= 1 else "")
        + (" * (1 + tax_rate)" if depth >= 2 else "")
        + f". Write the final materialized values in row order to Result!B2:B{n_rows+1}. "
        "Do not write formulas because grading reads cached values; ignore estimate and previous columns; save as output.xlsx."
    )
    return instruction, f"Result!B2:B{n_rows+1}", {"row_values": outputs}


def build_multi_step_pipeline(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    del task_id
    sales = wb.create_sheet("Sales")
    inventory = wb.create_sheet("Inventory")
    regions = wb.create_sheet("RegionMap")
    sku_key = ("sku", "item_code", "inventory_key")[ambiguity]
    store_key = ("store_id", "shop_id", "store_key")[ambiguity]
    sales.append(
        [
            "sale_id",
            sku_key,
            store_key,
            "units",
            "gross_amount",
            "status",
            "gross_amount_old" if ambiguity else "note",
        ]
    )
    inventory.append([sku_key, "category", "active_flag", "category_old" if ambiguity else "note"])
    regions.append([store_key, "region", "region_old" if ambiguity else "note"])
    skus = [(f"S{i:02d}", "Core" if i % 2 else "Other", 1 if i % 4 else 0) for i in range(1, 9)]
    stores = [(f"T{i:02d}", "East" if i % 2 else "West") for i in range(1, 7)]
    for key, category, active in skus:
        inventory.append([key, category, active, "Legacy" if ambiguity else "current"])
    for key, region in stores:
        regions.append([key, region, "North" if ambiguity else "current"])
    sku_meta = {key: (category, active) for key, category, active in skus}
    store_region = dict(stores)
    kept: list[tuple[int, float]] = []
    for index in range(16 + depth * 4):
        sku = skus[(index * 3 + 1) % len(skus)][0]
        store = stores[(index * 2 + 1) % len(stores)][0]
        units = 1 + index % 7
        gross = float(rng.randint(30, 240))
        status = "posted" if index % 5 else "void"
        sales.append([f"X{index+1:03d}", sku, store, units, gross, status, gross + 1000 if ambiguity else "live"])
        category, active = sku_meta[sku]
        keep = status == "posted"
        if depth >= 1:
            keep = keep and active == 1 and category == "Core"
        if depth >= 2:
            keep = keep and store_region[store] == "East"
        if keep:
            kept.append((units, gross))
    result = wb["Result"]
    result["B2"] = round(sum(gross for _, gross in kept), 2)
    result["B3"] = sum(units for units, _ in kept)
    result["B4"] = len(kept)
    instruction = "Start from Sales and keep only status posted. "
    if depth >= 1:
        instruction += f"Join Sales.{sku_key} to Inventory.{sku_key}, then keep active_flag 1 and category Core. "
    if depth >= 2:
        instruction += f"Join Sales.{store_key} to RegionMap.{store_key}, then keep region East. "
    instruction += (
        "Write the materialized sum of gross_amount to Result!B2, total units to Result!B3, and retained row count "
        "to Result!B4. Ignore old columns and save as output.xlsx."
    )
    return instruction, answer_range(4), {"kept": kept}
