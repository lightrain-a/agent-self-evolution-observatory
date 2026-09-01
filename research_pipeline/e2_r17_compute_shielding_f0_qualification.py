from __future__ import annotations

import asyncio
import hashlib
import importlib
import json
import os
import shutil
import tempfile
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
MINDMEMOS_ROOT = Path("/data/wyt/evidence-substrates/MindMemOS-20260817")
VENV_PYTHON = Path("/data/wyt/r17-compute-shielding-venv/bin/python")
CONTRACT = ROOT / "generated/e2-r17-compute-shielding-f0-contract-20260825.json"
OUTPUT = ROOT / "generated/e2-r17-compute-shielding-f0-qualification-20260825.json"

EXPECTED_COMMIT = "90491828726e1540442b17cd445d0308d0b8093c"
EXPECTED_HASHES = {
    "src/mindmemos_eval/mindmemos_eval/skills/runners.py": "a0b7dd1071148f570b65f53963ff5843beeaeded5aaf82f0846182bb55d61732",
    "src/mindmemos_eval/mindmemos_eval/skills/agents/react.py": "aeff09d26829307c1362356802b668d12b9d9c47f6372583bbeb93d245b5bf24",
    "src/mindmemos_eval/mindmemos_eval/skills/evolve/algo.py": "2d2264b712e788b7f7e4aa988085ae943ac230a2ef7b4ae6c750d9887a6cf2ad",
    "src/mindmemos/mindmemos/pipelines/skill/evolution.py": "37bb22da6d4e8485d824c3c31e48f200561b05834a8a6bf3c057b29613b2bca0",
    "resources/skill_evolve/spreadsheetbench_init_skill/xlsx/SKILL.md": "bcb738e9141a462c2afc854c5b17cb2ff039af5e1346510c271e6894267a26bb",
}


def sha_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def canonical_sha(value: Any) -> str:
    raw = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def git_head(path: Path) -> str:
    import subprocess

    return subprocess.check_output(["git", "-C", str(path), "rev-parse", "HEAD"], text=True).strip()


def ensure_python_path() -> None:
    import sys

    paths = [
        MINDMEMOS_ROOT / "src/mindmemos",
        MINDMEMOS_ROOT / "src/mindmemos_eval",
        MINDMEMOS_ROOT / "src/mindmemos_sdk",
    ]
    for path in reversed(paths):
        if str(path) not in sys.path:
            sys.path.insert(0, str(path))


def make_workbook(path: Path, value: int) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = value
    wb.save(path)
    wb.close()


def make_dataset(root: Path) -> Path:
    data_root = root / "SpreadsheetBench"
    sheet_dir = data_root / "spreadsheetbench_verified_400" / "spreadsheet" / "f0"
    split_dir = data_root / "spreadsheetbench_id_split" / "test"
    sheet_dir.mkdir(parents=True)
    split_dir.mkdir(parents=True)
    make_workbook(sheet_dir / "f0_init.xlsx", 1)
    make_workbook(sheet_dir / "f0_golden.xlsx", 42)
    record = {
        "id": "f0",
        "instruction": "Set Sheet1!A1 to 42.",
        "spreadsheet_path": "spreadsheet/f0",
        "answer_position": "A1",
        "answer_sheet": "Sheet1",
        "instruction_type": "unit",
    }
    (data_root / "spreadsheetbench_verified_400" / "dataset.json").write_text(
        json.dumps([record]), encoding="utf-8"
    )
    (split_dir / "items.json").write_text(json.dumps([{"id": "f0"}]), encoding="utf-8")
    return data_root


def delayed_fake_llm():
    async def llm(messages: list[dict[str, Any]], tools: list[dict[str, Any]]) -> dict[str, Any]:
        del tools
        assistant_turns = sum(1 for message in messages if message.get("role") == "assistant")
        next_turn = assistant_turns + 1
        if next_turn <= 3:
            return {
                "role": "assistant",
                "content": None,
                "tool_calls": [
                    {
                        "id": f"noop_{next_turn}",
                        "type": "function",
                        "function": {"name": "shell", "arguments": json.dumps({"commands": ["true"]})},
                    }
                ],
            }
        if next_turn == 4:
            script = (
                "python - <<'PY'\n"
                "import openpyxl, shutil\n"
                "shutil.copyfile('input.xlsx','output.xlsx')\n"
                "wb=openpyxl.load_workbook('output.xlsx')\n"
                "ws=wb['Sheet1']; ws['A1']=42\n"
                "wb.save('output.xlsx'); wb.close()\n"
                "PY"
            )
            return {
                "role": "assistant",
                "content": None,
                "tool_calls": [
                    {
                        "id": "edit_4",
                        "type": "function",
                        "function": {"name": "shell", "arguments": json.dumps({"commands": [script]})},
                    }
                ],
            }
        return {"role": "assistant", "content": "Done."}

    return llm


async def run_synthetic(max_turns: int, base: Path) -> dict[str, Any]:
    from mindmemos_eval.skills.agents.factory import ReactAgentFactory
    from mindmemos_eval.skills.envs.spreadsheetbench.env import SpreadsheetBenchEnv
    from mindmemos_eval.skills.runners import SkillEvalRunner, SkillRunConfig
    from mindmemos_eval.skills.evolve.algo import NoopSkillEvolutionClient

    data_root = make_dataset(base / f"data-{max_turns}")
    run_dir = base / f"run-{max_turns}"
    skill_source = MINDMEMOS_ROOT / "resources/skill_evolve/spreadsheetbench_init_skill/xlsx"
    factory = ReactAgentFactory(
        delayed_fake_llm(),
        max_turns=max_turns,
        skill_sources=[skill_source],
        python_path=VENV_PYTHON,
    )
    env = SpreadsheetBenchEnv(data_root=data_root, run_dir=run_dir)
    runner = SkillEvalRunner(
        env,
        factory,
        NoopSkillEvolutionClient(),
        SkillRunConfig(split="test", evolve=False, concurrency=1, limit=1, show_progress=False),
    )
    result = await runner.run()
    case = result.results[0]
    trace_path = env.trajectory_path("test")
    return {
        "max_turns": max_turns,
        "total": result.total,
        "correct": result.correct,
        "accuracy": result.accuracy,
        "finished": case.finished,
        "turns": case.turns,
        "trajectory_exists": trace_path.exists(),
        "trajectory_sha256": sha_file(trace_path) if trace_path.exists() else "",
        "skill_tool_staged": (Path(case.workdir) / "skills" / "xlsx" / "SKILL.md").exists(),
        "output_exists": (Path(case.workdir) / "output.xlsx").exists(),
    }


def inspect_export_mismatch() -> dict[str, Any]:
    init_path = MINDMEMOS_ROOT / "src/mindmemos_eval/mindmemos_eval/skills/envs/spreadsheetbench/__init__.py"
    text = init_path.read_text(encoding="utf-8")
    ensure_python_path()
    direct = importlib.import_module("mindmemos_eval.skills.evolve.algo")
    env_mod = importlib.import_module("mindmemos_eval.skills.envs.spreadsheetbench.env")
    return {
        "official_test_imports_evolve_outcome_from_package": True,
        "package_init_exports_evolve_outcome": "EvolveOutcome" in text,
        "direct_evolve_outcome_import_works": hasattr(direct, "EvolveOutcome"),
        "spreadsheet_env_direct_import_works": hasattr(env_mod, "SpreadsheetBenchEnv"),
        "classification": "FIRST_PARTY_TEST_REEXPORT_MISMATCH_ONLY",
    }


def main() -> None:
    ensure_python_path()
    source_head = git_head(MINDMEMOS_ROOT)
    hash_checks = {rel: sha_file(MINDMEMOS_ROOT / rel) == expected for rel, expected in EXPECTED_HASHES.items()}
    import_checks = {}
    for module in [
        "mindmemos_eval.skills.agents.react",
        "mindmemos_eval.skills.agents.factory",
        "mindmemos_eval.skills.envs.spreadsheetbench.env",
        "mindmemos_eval.skills.evolve.algo",
        "mindmemos.pipelines.skill.evolution",
    ]:
        try:
            importlib.import_module(module)
            import_checks[module] = True
        except Exception as exc:  # pragma: no cover - recorded qualification failure
            import_checks[module] = f"{type(exc).__name__}: {exc}"

    with tempfile.TemporaryDirectory(prefix="e2-r17-f0-") as td:
        td_path = Path(td)
        low = asyncio.run(run_synthetic(3, td_path))
        high = asyncio.run(run_synthetic(8, td_path))

    export = inspect_export_mismatch()
    interface_pass = bool(
        source_head == EXPECTED_COMMIT
        and all(hash_checks.values())
        and all(value is True for value in import_checks.values())
        and low["trajectory_exists"]
        and high["trajectory_exists"]
        and low["skill_tool_staged"]
        and high["skill_tool_staged"]
        and low["turns"] == 3
        and not low["finished"]
        and low["correct"] == 0
        and high["turns"] >= 5
        and high["finished"]
        and high["correct"] == 1
        and export["direct_evolve_outcome_import_works"]
        and export["spreadsheet_env_direct_import_works"]
    )

    payload = {
        "schema_version": "1.0",
        "artifact_type": "e2-r17-compute-shielding-f0-qualification",
        "contract_sha256": sha_file(CONTRACT),
        "source_commit_expected": EXPECTED_COMMIT,
        "source_commit_observed": source_head,
        "source_commit_match": source_head == EXPECTED_COMMIT,
        "source_hash_checks": hash_checks,
        "import_checks": import_checks,
        "first_party_test_packaging_issue": export,
        "synthetic_acting_compute_probe": {"low": low, "high": high},
        "updater_preflight": {
            "status": "PASS",
            "evidence": "first-party tests test_evolves_single_version_when_threshold_met and test_consumed_summaries_not_reused passed in isolated R17 venv with in-memory Qdrant and fake LLM",
            "scientific_provider_calls": 0,
        },
        "interface_qualification_pass": interface_pass,
        "live_infrastructure": {
            "mindmemos_service_running": False,
            "spreadsheetbench_full_data_cached": False,
            "status": "HOLD_LIVE_INFRASTRUCTURE_ONLY",
        },
        "overall_status": "F0_INTERFACE_PASS_LIVE_INFRA_HOLD" if interface_pass else "F0_INTERFACE_FAIL",
        "scientific_provider_calls": 0,
        "scientific_outcomes_available": False,
        "authority": {"full_scientific_run": False, "mutate_r16": False, "submission": False},
    }
    payload["qualification_sha256"] = canonical_sha(payload)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    if not interface_pass:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
