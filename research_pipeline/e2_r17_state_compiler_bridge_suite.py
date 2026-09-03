from __future__ import annotations

import hashlib
import json
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from research_pipeline.e2_r17_controlled_spreadsheet_suite import BUILDERS
from research_pipeline.e2_r17_controlled_suite_schema import (
    DISTRACTOR_COUNTS,
    FAMILIES,
    FAMILY_CODES,
    L9_PROFILES,
    BuiltTask,
    add_distractors,
    answer_cells,
    canonical_sha,
    new_book,
    normalize_xlsx,
    seeded_rng,
    select_by_hash,
    sha256_file,
    write_json,
)


BRIDGE_SUITE_ID = "E2-R17-STATE-COMPILER-BRIDGE-SUITE-V1"
BRIDGE_SCHEMA_VERSION = "1.0"
BRIDGE_BLOCK_ROLES = {
    7: "bridge_update_candidate",
    8: "bridge_update_candidate",
    9: "bridge_heldout_candidate",
}
SCREEN_SPLIT_SALT = "E2-R17-BRIDGE-SCREEN-VALIDATION-STREAM-v1"
UPDATE_SELECT_SALT = "E2-R17-BRIDGE-UPDATE-v1"
HELDOUT_SELECT_SALT = "E2-R17-BRIDGE-HELDOUT-v1"


def _build_task(verified_root: Path, *, block: int, family: str, profile_index: int) -> BuiltTask:
    if block not in BRIDGE_BLOCK_ROLES:
        raise ValueError(f"unknown bridge block: {block}")
    if family not in FAMILIES:
        raise ValueError(f"unknown family: {family}")
    if not 0 <= profile_index < len(L9_PROFILES):
        raise ValueError(f"unknown profile: {profile_index}")

    depth, distractor_level, ambiguity = L9_PROFILES[profile_index]
    task_id = f"r17-b{block}-{FAMILY_CODES[family]}-p{profile_index}"
    rng = seeded_rng(task_id)
    wb = new_book(task_id)
    distractor_names = add_distractors(wb, DISTRACTOR_COUNTS[distractor_level], rng, ambiguity)
    instruction, answer_position, expected = BUILDERS[family](wb, rng, depth, ambiguity, task_id)

    task_dir = verified_root / "spreadsheet" / task_id
    task_dir.mkdir(parents=True, exist_ok=True)
    init_path = task_dir / f"{task_id}_init.xlsx"
    golden_path = task_dir / f"{task_id}_golden.xlsx"
    expected_values = {
        f"{sheet}!{cell}": wb[sheet][cell].value
        for sheet, cell in answer_cells(answer_position)
    }
    for sheet, cell in answer_cells(answer_position):
        wb[sheet][cell] = None
    wb.save(init_path)
    normalize_xlsx(init_path)
    for qualified_cell, value in expected_values.items():
        sheet, cell = qualified_cell.split("!", 1)
        wb[sheet][cell] = value
    wb.save(golden_path)
    normalize_xlsx(golden_path)
    wb.close()

    record = {
        "id": task_id,
        "instruction": instruction,
        "spreadsheet_path": f"spreadsheet/{task_id}",
        "answer_position": answer_position,
        "answer_sheet": None,
        "instruction_type": family,
    }
    metadata = {
        "id": task_id,
        "suite_id": BRIDGE_SUITE_ID,
        "block": block,
        "role": BRIDGE_BLOCK_ROLES[block],
        "primary_failure_family": family,
        "profile_index": profile_index,
        "procedure_depth_level": depth,
        "distractor_level": distractor_level,
        "distractor_count": DISTRACTOR_COUNTS[distractor_level],
        "schema_ambiguity_level": ambiguity,
        "distractor_sheets": distractor_names,
        "answer_position": answer_position,
        "expected": expected,
        "golden_answer_cells": expected_values,
    }
    return BuiltTask(task_id, record, metadata, init_path, golden_path)


def _file_rows(root: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.name == "suite_manifest.json":
            continue
        rows.append(
            {
                "path": str(path.relative_to(root)),
                "size": path.stat().st_size,
                "sha256": sha256_file(path),
            }
        )
    return rows


def _make_update_streams(metadata: list[dict[str, Any]]) -> tuple[dict[str, list[str]], list[str]]:
    streams: dict[str, list[str]] = {}
    reserve: list[str] = []
    for family in FAMILIES:
        ids = sorted(
            row["id"]
            for row in metadata
            if row["role"] == "bridge_update_candidate"
            and row["primary_failure_family"] == family
        )
        if len(ids) != 18:
            raise AssertionError(f"bridge update candidate count drift for {family}: {len(ids)}")
        selected = select_by_hash(ids, count=16, salt=f"{UPDATE_SELECT_SALT}|{family}")
        reserve.extend(sorted(set(ids) - set(selected)))
        for index in range(2):
            stream_id = f"bridge-{FAMILY_CODES[family]}-{index:02d}"
            streams[stream_id] = selected[index * 8 : (index + 1) * 8]
    return streams, sorted(reserve)


def _split_streams(streams: dict[str, list[str]]) -> tuple[list[str], list[str]]:
    screen: list[str] = []
    validation: list[str] = []
    for family in FAMILIES:
        code = FAMILY_CODES[family]
        candidates = sorted(stream_id for stream_id in streams if f"-{code}-" in stream_id)
        if len(candidates) != 2:
            raise AssertionError(f"expected two bridge streams for {family}")
        ranked = sorted(
            candidates,
            key=lambda stream_id: hashlib.sha256(
                f"{SCREEN_SPLIT_SALT}|{stream_id}|{canonical_sha(streams[stream_id])}".encode("utf-8")
            ).hexdigest(),
        )
        screen.append(ranked[0])
        validation.append(ranked[1])
    return sorted(screen), sorted(validation)


def _make_heldout_panels(metadata: list[dict[str, Any]]) -> tuple[list[str], list[str], list[str]]:
    screen: list[str] = []
    validation: list[str] = []
    reserve: list[str] = []
    for family in FAMILIES:
        ids = sorted(
            row["id"]
            for row in metadata
            if row["role"] == "bridge_heldout_candidate"
            and row["primary_failure_family"] == family
        )
        if len(ids) != 9:
            raise AssertionError(f"bridge heldout candidate count drift for {family}: {len(ids)}")
        ranked = sorted(
            ids,
            key=lambda task_id: hashlib.sha256(
                f"{HELDOUT_SELECT_SALT}|{family}|{task_id}".encode("utf-8")
            ).hexdigest(),
        )
        screen.extend(sorted(ranked[:2]))
        validation.extend(sorted(ranked[2:4]))
        reserve.extend(sorted(ranked[4:]))
    return sorted(screen), sorted(validation), sorted(reserve)


def build_bridge_suite(output_root: Path, *, overwrite: bool = False) -> dict[str, Any]:
    if output_root.exists():
        if not overwrite:
            raise FileExistsError(output_root)
        shutil.rmtree(output_root)

    verified_root = output_root / "spreadsheetbench_verified_400"
    verified_root.mkdir(parents=True, exist_ok=True)
    records: list[dict[str, Any]] = []
    metadata: list[dict[str, Any]] = []

    for block in sorted(BRIDGE_BLOCK_ROLES):
        for family in FAMILIES:
            for profile_index in range(len(L9_PROFILES)):
                task = _build_task(verified_root, block=block, family=family, profile_index=profile_index)
                records.append(task.record)
                metadata.append(task.metadata)

    records.sort(key=lambda row: row["id"])
    metadata.sort(key=lambda row: row["id"])
    write_json(verified_root / "dataset.json", records)
    write_json(output_root / "bridge_metadata.json", metadata)

    streams, update_reserve = _make_update_streams(metadata)
    screen_streams, validation_streams = _split_streams(streams)
    screen_heldout, validation_heldout, heldout_reserve = _make_heldout_panels(metadata)

    split = {
        "schema_version": BRIDGE_SCHEMA_VERSION,
        "suite_id": BRIDGE_SUITE_ID,
        "selection_is_outcome_blind": True,
        "selection_algorithm": "content-addressed SHA-256 salts only",
        "update_streams": streams,
        "screen_stream_ids": screen_streams,
        "validation_stream_ids": validation_streams,
        "update_reserve_integrity_only": update_reserve,
        "screen_heldout": screen_heldout,
        "validation_heldout": validation_heldout,
        "heldout_reserve_integrity_only": heldout_reserve,
        "rules": {
            "screen_validation_streams_family_balanced": True,
            "screen_validation_heldout_family_balanced": True,
            "screen_validation_heldout_disjoint": True,
            "heldout_never_fed_to_updater": True,
            "validation_never_opened_before_screen_gate": True,
            "reserve_only_for_preexecution_file_integrity_failure": True,
            "reserve_never_replaces_model_failure_or_bad_outcome": True,
            "e3_blocks_5_6_not_used": True,
        },
    }
    write_json(output_root / "bridge_split_manifest.json", split)

    rows = _file_rows(output_root)
    manifest = {
        "schema_version": BRIDGE_SCHEMA_VERSION,
        "suite_id": BRIDGE_SUITE_ID,
        "candidate_task_count": len(records),
        "formal_task_count": 96 + 12 + 12,
        "update_task_count": 96,
        "screen_heldout_task_count": 12,
        "validation_heldout_task_count": 12,
        "update_stream_count": 12,
        "screen_stream_count": 6,
        "validation_stream_count": 6,
        "families": list(FAMILIES),
        "blocks": {str(k): v for k, v in BRIDGE_BLOCK_ROLES.items()},
        "dataset_sha256": canonical_sha(rows),
        "files": rows,
        "split_manifest_sha256": sha256_file(output_root / "bridge_split_manifest.json"),
        "metadata_sha256": sha256_file(output_root / "bridge_metadata.json"),
        "dataset_json_sha256": sha256_file(verified_root / "dataset.json"),
    }
    write_json(output_root / "suite_manifest.json", manifest)
    return manifest


def self_check_bridge_suite(output_root: Path) -> dict[str, Any]:
    records = json.loads((output_root / "spreadsheetbench_verified_400/dataset.json").read_text(encoding="utf-8"))
    metadata = json.loads((output_root / "bridge_metadata.json").read_text(encoding="utf-8"))
    split = json.loads((output_root / "bridge_split_manifest.json").read_text(encoding="utf-8"))
    manifest = json.loads((output_root / "suite_manifest.json").read_text(encoding="utf-8"))

    ids = [row["id"] for row in records]
    if len(ids) != 162 or len(ids) != len(set(ids)) or len(metadata) != 162:
        raise AssertionError("bridge candidate cardinality mismatch")
    if any(not task_id.startswith(("r17-b7-", "r17-b8-", "r17-b9-")) for task_id in ids):
        raise AssertionError("bridge task escaped blocks 7-9")

    streams = split["update_streams"]
    if len(streams) != 12 or any(len(task_ids) != 8 for task_ids in streams.values()):
        raise AssertionError("bridge stream shape mismatch")
    update_ids = {task_id for task_ids in streams.values() for task_id in task_ids}
    if len(update_ids) != 96:
        raise AssertionError("bridge update task count mismatch")
    if len(split["update_reserve_integrity_only"]) != 12:
        raise AssertionError("bridge update reserve mismatch")

    screen = set(split["screen_heldout"])
    validation = set(split["validation_heldout"])
    if len(screen) != 12 or len(validation) != 12 or screen & validation:
        raise AssertionError("screen/validation heldout disjointness failure")
    if update_ids & (screen | validation):
        raise AssertionError("heldout leaked into update tasks")
    if len(split["heldout_reserve_integrity_only"]) != 30:
        raise AssertionError("bridge heldout reserve mismatch")

    meta_by_id = {row["id"]: row for row in metadata}
    for stream_id, task_ids in streams.items():
        families = {meta_by_id[x]["primary_failure_family"] for x in task_ids}
        if len(families) != 1:
            raise AssertionError(f"non-homogeneous bridge stream: {stream_id}")

    for panel_name, panel in (("screen", screen), ("validation", validation)):
        counts = Counter(meta_by_id[x]["primary_failure_family"] for x in panel)
        if counts != Counter({family: 2 for family in FAMILIES}):
            raise AssertionError(f"{panel_name} family balance mismatch: {counts}")

    for group_name in ("screen_stream_ids", "validation_stream_ids"):
        counts = Counter(
            meta_by_id[streams[stream_id][0]]["primary_failure_family"]
            for stream_id in split[group_name]
        )
        if counts != Counter({family: 1 for family in FAMILIES}):
            raise AssertionError(f"{group_name} family balance mismatch: {counts}")

    for record in records:
        task_dir = output_root / "spreadsheetbench_verified_400" / record["spreadsheet_path"]
        init = next(task_dir.glob("*init*.xlsx"))
        golden = next(task_dir.glob("*golden*.xlsx"))
        wb_init = load_workbook(init, data_only=True)
        wb_gold = load_workbook(golden, data_only=True)
        try:
            for sheet, cell in answer_cells(record["answer_position"]):
                if wb_init[sheet][cell].value is not None:
                    raise AssertionError(f"bridge init answer not blank: {record['id']} {sheet}!{cell}")
                if wb_gold[sheet][cell].value is None:
                    raise AssertionError(f"bridge golden answer blank: {record['id']} {sheet}!{cell}")
        finally:
            wb_init.close()
            wb_gold.close()

    observed_sha = canonical_sha(_file_rows(output_root))
    if observed_sha != manifest["dataset_sha256"]:
        raise AssertionError("bridge dataset SHA mismatch")

    return {
        "status": "PASS",
        "candidate_task_count": 162,
        "formal_task_count": 120,
        "update_task_count": 96,
        "screen_heldout_task_count": 12,
        "validation_heldout_task_count": 12,
        "screen_validation_heldout_disjoint": True,
        "screen_stream_count": 6,
        "validation_stream_count": 6,
        "dataset_sha256": observed_sha,
    }


__all__ = [
    "BRIDGE_SUITE_ID",
    "BRIDGE_BLOCK_ROLES",
    "build_bridge_suite",
    "self_check_bridge_suite",
]
