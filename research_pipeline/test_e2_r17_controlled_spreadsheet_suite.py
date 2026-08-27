from __future__ import annotations

import json
import tempfile
import unittest
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from research_pipeline.e2_r17_controlled_spreadsheet_suite import build_suite, self_check_suite
from research_pipeline.e2_r17_controlled_suite_schema import FAMILIES, L9_PROFILES, answer_cells


class ControlledSpreadsheetSuiteTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.temp = tempfile.TemporaryDirectory()
        cls.root = Path(cls.temp.name) / "suite"
        cls.manifest = build_suite(cls.root)
        cls.check = self_check_suite(cls.root)
        cls.records = json.loads(
            (cls.root / "spreadsheetbench_verified_400/dataset.json").read_text(encoding="utf-8")
        )
        cls.metadata = json.loads((cls.root / "r17_controlled_metadata.json").read_text(encoding="utf-8"))
        cls.split = json.loads((cls.root / "r17_split_manifest.json").read_text(encoding="utf-8"))

    @classmethod
    def tearDownClass(cls) -> None:
        cls.temp.cleanup()

    def test_cardinality_and_families(self) -> None:
        self.assertEqual(len(self.records), 378)
        self.assertEqual(len({row["id"] for row in self.records}), 378)
        self.assertEqual(set(row["instruction_type"] for row in self.records), set(FAMILIES))
        self.assertEqual(self.check["status"], "PASS")

    def test_l9_is_pairwise_balanced(self) -> None:
        for column in range(3):
            self.assertEqual(Counter(row[column] for row in L9_PROFILES), {0: 3, 1: 3, 2: 3})
        for left, right in ((0, 1), (0, 2), (1, 2)):
            pairs = Counter((row[left], row[right]) for row in L9_PROFILES)
            self.assertEqual(set(pairs.values()), {1})
            self.assertEqual(len(pairs), 9)

    def test_split_shape_and_disjointness(self) -> None:
        streams = self.split["e1_update_streams"]
        self.assertEqual(len(streams), 12)
        self.assertTrue(all(len(ids) == 8 for ids in streams.values()))
        future_streams = self.split["e3_future_streams"]
        self.assertEqual(len(future_streams), 12)
        self.assertTrue(all(len(ids) == 8 for ids in future_streams.values()))
        update = {task_id for ids in streams.values() for task_id in ids}
        future = {task_id for ids in future_streams.values() for task_id in ids}
        groups = [
            set(self.split["development"]),
            set(self.split["e0_calibration"]),
            update,
            set(self.split["e1_update_reserve_integrity_only"]),
            set(self.split["e1_common_heldout_probe"]),
            future,
            set(self.split["e3_future_reserve_integrity_only"]),
        ]
        for index, left in enumerate(groups):
            for right in groups[index + 1 :]:
                self.assertFalse(left & right)

    def test_streams_are_family_homogeneous(self) -> None:
        meta = {row["id"]: row for row in self.metadata}
        for key in ("e1_update_streams", "e3_future_streams"):
            for task_ids in self.split[key].values():
                self.assertEqual(len({meta[task_id]["primary_failure_family"] for task_id in task_ids}), 1)

    def test_probe_is_family_balanced(self) -> None:
        meta = {row["id"]: row for row in self.metadata}
        counts = Counter(meta[task_id]["primary_failure_family"] for task_id in self.split["e1_common_heldout_probe"])
        self.assertEqual(counts, Counter({family: 3 for family in FAMILIES}))

    def test_input_answers_are_blank_and_golden_answers_materialized(self) -> None:
        for record in self.records[::29]:
            task_dir = self.root / "spreadsheetbench_verified_400" / record["spreadsheet_path"]
            init = next(task_dir.glob("*init*.xlsx"))
            golden = next(task_dir.glob("*golden*.xlsx"))
            wb_init = load_workbook(init, data_only=True)
            wb_gold = load_workbook(golden, data_only=True)
            try:
                for sheet, cell in answer_cells(record["answer_position"]):
                    self.assertIsNone(wb_init[sheet][cell].value)
                    self.assertIsNotNone(wb_gold[sheet][cell].value)
            finally:
                wb_init.close()
                wb_gold.close()

    def test_rebuild_is_byte_identical_at_dataset_manifest_level(self) -> None:
        other = Path(self.temp.name) / "suite2"
        manifest2 = build_suite(other)
        self.assertEqual(self.manifest["dataset_sha256"], manifest2["dataset_sha256"])


if __name__ == "__main__":
    unittest.main()
