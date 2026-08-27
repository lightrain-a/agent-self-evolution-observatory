from __future__ import annotations

import hashlib
import json
import random
import re
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

SCHEMA_VERSION = "1.0"
SUITE_ID = "E2-R17-CONTROLLED-SPREADSHEET-SUITE-V1"
FAMILIES = (
    "input_output_contract",
    "target_sheet_range",
    "schema_key_alignment",
    "aggregation_join",
    "formula_materialization",
    "multi_step_pipeline",
)
FAMILY_CODES = {
    "input_output_contract": "ioc",
    "target_sheet_range": "tsr",
    "schema_key_alignment": "ska",
    "aggregation_join": "agj",
    "formula_materialization": "fmv",
    "multi_step_pipeline": "msp",
}
# Orthogonal L9 design. Each factor level occurs three times and every pair of
# factor levels occurs exactly once.
L9_PROFILES = (
    (0, 0, 0),
    (0, 1, 1),
    (0, 2, 2),
    (1, 0, 1),
    (1, 1, 2),
    (1, 2, 0),
    (2, 0, 2),
    (2, 1, 0),
    (2, 2, 1),
)
BLOCK_ROLES = {
    0: "development",
    1: "e0_calibration",
    2: "e1_update_candidate",
    3: "e1_update_candidate",
    4: "e1_heldout_probe_candidate",
    5: "e3_future_candidate",
}
DISTRACTOR_COUNTS = {0: 0, 1: 2, 2: 5}


@dataclass(frozen=True)
class BuiltTask:
    task_id: str
    record: dict[str, Any]
    metadata: dict[str, Any]
    init_path: Path
    golden_path: Path


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_sha(payload: Any) -> str:
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def seeded_rng(task_id: str) -> random.Random:
    seed = int(hashlib.sha256(task_id.encode("utf-8")).hexdigest()[:16], 16)
    return random.Random(seed)


def new_book(task_id: str) -> Workbook:
    wb = Workbook()
    fixed_time = datetime(2000, 1, 1, 0, 0, 0)
    wb.properties.created = fixed_time
    wb.properties.modified = fixed_time
    wb.properties.creator = "E2-R17 controlled-suite generator"
    wb.properties.lastModifiedBy = "E2-R17 controlled-suite generator"
    ws = wb.active
    ws.title = "Result"
    ws["A1"] = "E2-R17 controlled result"
    ws["A2"] = "metric_1"
    ws["A3"] = "metric_2"
    ws["A4"] = "metric_3"
    ws["A5"] = "metric_4"
    ws["D1"] = "immutable_task_sentinel"
    ws["E1"] = task_id
    return wb


def add_distractors(wb: Workbook, count: int, rng: random.Random, ambiguity: int) -> list[str]:
    names: list[str] = []
    for index in range(count):
        suffix = chr(ord("A") + index)
        name = f"Archive_{suffix}" if ambiguity < 2 else f"Data_{suffix}_candidate"
        ws = wb.create_sheet(name)
        ws.append(["code", "amount", "status", "note"])
        for row in range(1, 8):
            ws.append(
                [
                    f"D{index}{row}",
                    rng.randint(10, 500),
                    "inactive" if row % 2 else "active",
                    f"distractor-{index}-{row}",
                ]
            )
        names.append(name)
    return names


def answer_range(last_row: int = 4) -> str:
    return f"Result!B2:B{last_row}"


def normalize_xlsx(path: Path) -> None:
    """Rewrite an XLSX ZIP with deterministic entry order and metadata."""
    with zipfile.ZipFile(path, "r") as source:
        entries = [(info.filename, source.read(info.filename), info.compress_type) for info in source.infolist()]
    with tempfile.NamedTemporaryFile(dir=path.parent, suffix=".xlsx", delete=False) as handle:
        temp_path = Path(handle.name)
    try:
        with zipfile.ZipFile(temp_path, "w") as target:
            for name, payload, compress_type in sorted(entries, key=lambda row: row[0]):
                if name == "docProps/core.xml":
                    payload = re.sub(
                        br"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                        b'<dcterms:modified xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>',
                        payload,
                    )
                info = zipfile.ZipInfo(filename=name, date_time=(1980, 1, 1, 0, 0, 0))
                info.compress_type = compress_type
                info.create_system = 3
                info.external_attr = 0o600 << 16
                target.writestr(info, payload)
        temp_path.replace(path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def answer_cells(answer_position: str) -> list[tuple[str, str]]:
    cells: list[tuple[str, str]] = []
    for segment in answer_position.split(","):
        sheet_name, cell_range = segment.strip().split("!", 1)
        if ":" not in cell_range:
            cells.append((sheet_name, cell_range))
            continue
        start, end = cell_range.split(":", 1)
        start_col = "".join(char for char in start if char.isalpha())
        end_col = "".join(char for char in end if char.isalpha())
        start_row = int("".join(char for char in start if char.isdigit()))
        end_row = int("".join(char for char in end if char.isdigit()))
        if start_col != end_col:
            raise ValueError("controlled suite currently supports one-column answer ranges")
        cells.extend((sheet_name, f"{start_col}{row}") for row in range(start_row, end_row + 1))
    return cells


def select_by_hash(ids: list[str], *, count: int, salt: str) -> list[str]:
    return sorted(ids, key=lambda value: hashlib.sha256(f"{salt}|{value}".encode()).hexdigest())[:count]
