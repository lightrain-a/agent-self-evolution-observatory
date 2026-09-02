from __future__ import annotations

import random
from typing import Any

from openpyxl import Workbook

from .e2_r17_controlled_suite_schema import answer_range


SEMANTIC_TYPES = ("PROCEDURAL_TRANSFORMATION", "INSTANCE_BINDING_LOCALIZATION")

FAMILY_SPECS: dict[str, dict[str, Any]] = {
    "ordered_filter_rollup": {
        "code": "ofr",
        "semantic_type": "PROCEDURAL_TRANSFORMATION",
        "matched_skeleton": "two_table_join",
        "reusable_transform_steps": 3,
        "binding_candidate_count": 1,
    },
    "foreign_key_binding": {
        "code": "fkb",
        "semantic_type": "INSTANCE_BINDING_LOCALIZATION",
        "matched_skeleton": "two_table_join",
        "reusable_transform_steps": 1,
        "binding_candidate_count": 3,
    },
    "normalize_then_rank": {
        "code": "ntr",
        "semantic_type": "PROCEDURAL_TRANSFORMATION",
        "matched_skeleton": "single_table_measure",
        "reusable_transform_steps": 3,
        "binding_candidate_count": 1,
    },
    "header_source_binding": {
        "code": "hsb",
        "semantic_type": "INSTANCE_BINDING_LOCALIZATION",
        "matched_skeleton": "single_table_measure",
        "reusable_transform_steps": 1,
        "binding_candidate_count": 3,
    },
    "reconcile_then_aggregate": {
        "code": "rta",
        "semantic_type": "PROCEDURAL_TRANSFORMATION",
        "matched_skeleton": "snapshot_table",
        "reusable_transform_steps": 3,
        "binding_candidate_count": 1,
    },
    "named_region_binding": {
        "code": "nrb",
        "semantic_type": "INSTANCE_BINDING_LOCALIZATION",
        "matched_skeleton": "snapshot_table",
        "reusable_transform_steps": 1,
        "binding_candidate_count": 3,
    },
}

FAMILIES = tuple(FAMILY_SPECS)
FAMILY_CODES = {name: str(spec["code"]) for name, spec in FAMILY_SPECS.items()}


def _semantic_expected(family: str, payload: dict[str, Any]) -> dict[str, Any]:
    spec = FAMILY_SPECS[family]
    return {
        **payload,
        "semantic_type": spec["semantic_type"],
        "matched_skeleton": spec["matched_skeleton"],
        "reusable_transform_steps": spec["reusable_transform_steps"],
        "binding_candidate_count": spec["binding_candidate_count"],
    }


def build_ordered_filter_rollup(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    del task_id
    orders = wb.create_sheet("Orders")
    accounts = wb.create_sheet("Accounts")
    orders.append(["order_id", "account_id", "units", "unit_price", "discount_rate", "status", "unit_price_old"])
    accounts.append(["account_id", "segment", "segment_previous"])
    account_rows = [(f"A{i:02d}", "Core" if i % 2 else "Other") for i in range(1, 9)]
    for key, segment in account_rows:
        accounts.append([key, segment, "Legacy" if ambiguity else "previous"])
    segment_by_account = dict(account_rows)
    retained: list[float] = []
    for i in range(14 + depth * 4):
        account = account_rows[(i * 3 + 1) % len(account_rows)][0]
        units = 1 + i % 5
        price = rng.randint(15, 95)
        discount = (i % 3) * 0.05
        status = "posted" if i % 4 else "void"
        orders.append([f"O{i+1:03d}", account, units, price, discount, status, price + 500])
        keep = status == "posted"
        if depth >= 1:
            keep = keep and segment_by_account[account] == "Core"
        value = float(units * price)
        if depth >= 2:
            value *= 1.0 - discount
        if keep:
            retained.append(round(value, 2))
    result = wb["Result"]
    result["B2"] = round(sum(retained), 2)
    result["B3"] = len(retained)
    result["B4"] = round(sum(retained) / len(retained), 2) if retained else 0.0
    instruction = "Start from Orders, keep only status posted, and compute units * unit_price for retained rows. "
    if depth >= 1:
        instruction += "Join Orders.account_id to Accounts.account_id and additionally keep only segment Core. "
    if depth >= 2:
        instruction += "Before aggregation, apply each retained row's discount_rate as value * (1 - discount_rate). "
    instruction += (
        "Write the materialized retained-value sum to Result!B2, retained row count to Result!B3, and retained-value mean "
        "rounded to 2 decimals to Result!B4. Ignore old/previous columns and save as output.xlsx."
    )
    return instruction, answer_range(4), _semantic_expected("ordered_filter_rollup", {"retained": retained})


def build_foreign_key_binding(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    del task_id
    ledger = wb.create_sheet("Ledger")
    mapping = wb.create_sheet("AccountMap")
    current_left = ("account_id", "account_code", "ledger_account_key")[ambiguity]
    current_right = ("account_id", "rate_account_code", "current_account_key")[ambiguity]
    legacy_left = ("legacy_id", "legacy_account_id", "account_key_previous")[ambiguity]
    candidate_left = ("candidate_id", "candidate_account_code", "account_key_candidate")[ambiguity]
    legacy_right = ("legacy_map_id", "legacy_rate_account", "rate_key_previous")[ambiguity]
    candidate_right = ("candidate_map_id", "candidate_rate_account", "rate_key_candidate")[ambiguity]
    left_candidates = [current_left, legacy_left, candidate_left]
    right_candidates = [current_right, legacy_right, candidate_right]
    rng.shuffle(left_candidates)
    rng.shuffle(right_candidates)
    ledger.append(["row_id", *left_candidates, "amount", "status"])
    mapping.append([*right_candidates, "multiplier"])
    keys = [f"K{i:02d}" for i in range(1, 7)]
    multipliers = {key: round(0.8 + i * 0.07, 2) for i, key in enumerate(keys)}
    for i, key in enumerate(keys):
        values = {
            current_right: key,
            legacy_right: f"MAP-L{i:02d}",
            candidate_right: f"MAP-C{i:02d}",
        }
        mapping.append([*(values[h] for h in right_candidates), multipliers[key]])
    converted: list[float] = []
    for i in range(10 + depth * 3):
        key = keys[i % len(keys)]
        amount = rng.randint(30, 220)
        status = "active" if i % 4 else "inactive"
        values = {
            current_left: key,
            legacy_left: f"LED-L{(i+2)%len(keys):02d}",
            candidate_left: f"LED-C{(i+3)%len(keys):02d}",
        }
        ledger.append([f"R{i+1:03d}", *(values[h] for h in left_candidates), amount, status])
        if depth == 0 or status == "active":
            converted.append(round(amount * multipliers[key], 2))
    result = wb["Result"]
    result["B2"] = round(sum(converted), 2)
    result["B3"] = len(converted)
    last = 3
    if depth >= 2:
        result["B4"] = max(converted) if converted else 0.0
        last = 4
    instruction = (
        "Identify the authoritative current account identifier in Ledger and bind it to the authoritative current account key "
        "in AccountMap; do not use legacy, previous, or candidate key columns. Multiply Ledger.amount by the matched "
        "AccountMap.multiplier. "
    )
    if depth >= 1:
        instruction += "Use only Ledger rows whose status is active. "
    instruction += "Write the materialized total to Result!B2 and retained row count to Result!B3. "
    if depth >= 2:
        instruction += "Write the maximum converted row amount to Result!B4. "
    instruction += "Save as output.xlsx."
    return instruction, answer_range(last), _semantic_expected(
        "foreign_key_binding",
        {
            "left_key": current_left,
            "right_key": current_right,
            "left_candidate_order": left_candidates,
            "right_candidate_order": right_candidates,
            "converted": converted,
        },
    )


def build_normalize_then_rank(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    del task_id, ambiguity
    ws = wb.create_sheet("Measures")
    ws.append(["item_id", "raw_amount", "scale", "active_flag", "raw_amount_old"])
    normalized: list[float] = []
    for i in range(9 + depth * 3):
        amount = rng.randint(80, 900)
        scale = (1, 10, 100)[i % 3]
        active = 0 if i % 5 == 0 else 1
        ws.append([f"M{i+1:02d}", amount, scale, active, amount + 1000])
        value = round(amount / scale, 2)
        if depth == 0 or active == 1:
            normalized.append(value)
    selected = list(normalized)
    if depth >= 2:
        selected = sorted(selected, reverse=True)[:3]
    result = wb["Result"]
    result["B2"] = round(sum(selected), 2)
    result["B3"] = max(selected) if selected else 0.0
    result["B4"] = len(selected)
    instruction = "For every Measures row, normalize raw_amount by dividing by scale. "
    if depth >= 1:
        instruction += "Then retain only rows with active_flag = 1. "
    if depth >= 2:
        instruction += "Then keep the three largest normalized values. "
    instruction += (
        "Write the materialized selected-value sum to Result!B2, maximum to Result!B3, and selected count to Result!B4. "
        "Ignore raw_amount_old and save as output.xlsx."
    )
    return instruction, answer_range(4), _semantic_expected("normalize_then_rank", {"selected": selected})


def build_header_source_binding(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    del task_id
    ws = wb.create_sheet("Metrics")
    authoritative = ("amount", "posted_amount", "recognized_amount")[ambiguity]
    estimate = ("amount_estimate", "posted_amount_estimate", "recognized_amount_estimate")[ambiguity]
    previous = ("amount_previous", "posted_amount_previous", "recognized_amount_previous")[ambiguity]
    amount_candidates = [authoritative, estimate, previous]
    rng.shuffle(amount_candidates)
    ws.append(["row_id", *amount_candidates, "status"])
    values: list[float] = []
    for i in range(10 + depth * 2):
        value = float(rng.randint(15, 180))
        status = "active" if i % 4 else "inactive"
        fields = {authoritative: value, estimate: value + 500, previous: value + 900}
        ws.append([f"H{i+1:02d}", *(fields[h] for h in amount_candidates), status])
        if depth == 0 or status == "active":
            values.append(value)
    result = wb["Result"]
    result["B2"] = round(sum(values), 2)
    result["B3"] = round(sum(values) / len(values), 2) if values else 0.0
    last = 3
    if depth >= 2:
        result["B4"] = sum(1 for value in values if value >= 100)
        last = 4
    instruction = (
        "Use the authoritative current amount field in Metrics, not its estimate or previous counterpart. "
    )
    if depth >= 1:
        instruction += "Use only rows whose status is active. "
    instruction += "Write the materialized amount sum to Result!B2 and arithmetic mean rounded to 2 decimals to Result!B3. "
    if depth >= 2:
        instruction += "Write the count of selected amounts at least 100 to Result!B4. "
    instruction += "Save as output.xlsx."
    return instruction, answer_range(last), _semantic_expected(
        "header_source_binding",
        {"authoritative_header": authoritative, "candidate_order": amount_candidates, "values": values},
    )


def build_reconcile_then_aggregate(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    del task_id
    snapshot = wb.create_sheet("Snapshot")
    corrections = wb.create_sheet("Corrections")
    snapshot.append(["record_id", "amount", "status", "amount_old"])
    corrections.append(["record_id", "replacement_amount", "replacement_status", "apply_flag", "replacement_amount_old"])
    state: dict[str, tuple[float, str]] = {}
    n = 10 + depth * 2
    for i in range(n):
        rid = f"S{i+1:02d}"
        amount = float(rng.randint(25, 210))
        status = "active" if i % 4 else "inactive"
        snapshot.append([rid, amount, status, amount + 800])
        state[rid] = (amount, status)
    for i in range(0, n, 3):
        rid = f"S{i+1:02d}"
        amount, status = state[rid]
        replacement_amount = amount + 10 + (i % 5)
        replacement_status = "active" if i % 2 == 0 else status
        apply_flag = 1 if i % 6 != 3 else 0
        corrections.append([rid, replacement_amount, replacement_status, apply_flag, replacement_amount + 900])
        if apply_flag == 1:
            state[rid] = (replacement_amount, replacement_status)
    values = list(state.values())
    if depth >= 1:
        values = [row for row in values if row[1] == "active"]
    amounts = [row[0] for row in values]
    if depth >= 2:
        amounts = [value for value in amounts if value >= 80]
    result = wb["Result"]
    result["B2"] = round(sum(amounts), 2)
    result["B3"] = len(amounts)
    result["B4"] = max(amounts) if amounts else 0.0
    instruction = (
        "Start from Snapshot. Apply Corrections rows with apply_flag = 1 by record_id, replacing both amount and status; "
        "ignore correction rows with apply_flag = 0. "
    )
    if depth >= 1:
        instruction += "After reconciliation, retain only active records. "
    if depth >= 2:
        instruction += "Then retain only reconciled amounts at least 80. "
    instruction += (
        "Write the materialized reconciled amount sum to Result!B2, retained record count to Result!B3, and retained maximum "
        "to Result!B4. Ignore old columns and save as output.xlsx."
    )
    return instruction, answer_range(4), _semantic_expected("reconcile_then_aggregate", {"amounts": amounts})


def build_named_region_binding(
    wb: Workbook, rng: random.Random, depth: int, ambiguity: int, task_id: str
) -> tuple[str, str, dict[str, Any]]:
    del task_id
    ws = wb.create_sheet("SnapshotBundle")
    starts = [1, 5, 9]
    if ambiguity == 0:
        region_defs = [
            ("CURRENT_ACTUAL", True),
            ("FORECAST_CANDIDATE", False),
            ("ARCHIVE_PREVIOUS", False),
        ]
    elif ambiguity == 1:
        region_defs = [
            ("AUTHORITATIVE_CURRENT", True),
            ("CURRENT_FORECAST", False),
            ("PREVIOUS_ACTUAL", False),
        ]
    else:
        region_defs = [
            ("FINAL_CURRENT_ACTUAL", True),
            ("FINAL_CANDIDATE", False),
            ("FINAL_PREVIOUS", False),
        ]
    rng.shuffle(region_defs)
    target_values: list[float] = []
    target_label = ""
    for idx, (label, is_target) in enumerate(region_defs):
        start = starts[idx]
        ws.cell(row=1, column=start, value=label)
        ws.cell(row=2, column=start, value="item_id")
        ws.cell(row=2, column=start + 1, value="amount")
        ws.cell(row=2, column=start + 2, value="status")
        values: list[float] = []
        for r in range(7 + depth):
            amount = float(rng.randint(20, 160) + idx * 400)
            status = "active" if r % 3 else "inactive"
            ws.cell(row=3 + r, column=start, value=f"B{idx}{r:02d}")
            ws.cell(row=3 + r, column=start + 1, value=amount)
            ws.cell(row=3 + r, column=start + 2, value=status)
            if depth == 0 or status == "active":
                values.append(amount)
        if is_target:
            target_values = values
            target_label = label
    result = wb["Result"]
    result["B2"] = round(sum(target_values), 2)
    result["B3"] = max(target_values) if target_values else 0.0
    last = 3
    if depth >= 2:
        result["B4"] = len(target_values)
        last = 4
    instruction = (
        "In SnapshotBundle, locate the authoritative current/actual region; do not use forecast, candidate, archive, or previous "
        "regions. Use that region's amount column. "
    )
    if depth >= 1:
        instruction += "Use only rows in that region whose status is active. "
    instruction += "Write the materialized amount sum to Result!B2 and maximum to Result!B3. "
    if depth >= 2:
        instruction += "Write the retained row count to Result!B4. "
    instruction += "Save as output.xlsx."
    return instruction, answer_range(last), _semantic_expected(
        "named_region_binding", {"region_label": target_label, "region_order": [x[0] for x in region_defs], "target_values": target_values}
    )


BUILDERS = {
    "ordered_filter_rollup": build_ordered_filter_rollup,
    "foreign_key_binding": build_foreign_key_binding,
    "normalize_then_rank": build_normalize_then_rank,
    "header_source_binding": build_header_source_binding,
    "reconcile_then_aggregate": build_reconcile_then_aggregate,
    "named_region_binding": build_named_region_binding,
}
