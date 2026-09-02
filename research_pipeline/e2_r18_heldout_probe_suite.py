from __future__ import annotations
import hashlib, json, shutil
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from .e2_r17_controlled_spreadsheet_suite import BUILDERS
from .e2_r17_controlled_suite_schema import (
    DISTRACTOR_COUNTS,FAMILIES,FAMILY_CODES,L9_PROFILES,BuiltTask,add_distractors,
    answer_cells,new_book,normalize_xlsx,seeded_rng,sha256_file,write_json
)

SUITE_ID='E2-R18-FRESH-HELDOUT-PROBE-SUITE-V1'
BLOCK=7
ROLE='r18_heldout_probe_candidate'
SALT='r18-heldout-v1'

def task_id(family:str, profile_index:int)->str:
    return f'r18-b{BLOCK}-{FAMILY_CODES[family]}-p{profile_index}'

def build_task(verified_root:Path, *, family:str, profile_index:int)->BuiltTask:
    depth,distractor_level,ambiguity=L9_PROFILES[profile_index]
    tid=task_id(family,profile_index)
    rng=seeded_rng(tid); wb=new_book(tid)
    distractor_names=add_distractors(wb,DISTRACTOR_COUNTS[distractor_level],rng,ambiguity)
    instruction,answer_position,expected=BUILDERS[family](wb,rng,depth,ambiguity,tid)
    td=verified_root/'spreadsheet'/tid; td.mkdir(parents=True,exist_ok=True)
    init=td/f'{tid}_init.xlsx'; golden=td/f'{tid}_golden.xlsx'
    expected_values={f'{s}!{c}':wb[s][c].value for s,c in answer_cells(answer_position)}
    for s,c in answer_cells(answer_position): wb[s][c]=None
    wb.save(init); normalize_xlsx(init)
    for q,v in expected_values.items():
        s,c=q.split('!',1); wb[s][c]=v
    wb.save(golden); normalize_xlsx(golden); wb.close()
    record={'id':tid,'instruction':instruction,'spreadsheet_path':f'spreadsheet/{tid}','answer_position':answer_position,'answer_sheet':None,'instruction_type':family}
    meta={'id':tid,'suite_id':SUITE_ID,'block':BLOCK,'role':ROLE,'primary_failure_family':family,'profile_index':profile_index,'procedure_depth_level':depth,'distractor_level':distractor_level,'distractor_count':DISTRACTOR_COUNTS[distractor_level],'schema_ambiguity_level':ambiguity,'distractor_sheets':distractor_names,'answer_position':answer_position,'expected':expected,'golden_answer_cells':expected_values}
    return BuiltTask(tid,record,meta,init,golden)

def select(ids:list[str],count:int,salt:str)->list[str]:
    return sorted(ids,key=lambda x: hashlib.sha256(f'{salt}|{x}'.encode()).hexdigest())[:count]

def file_rows(root:Path)->list[dict[str,Any]]:
    rows=[]
    for p in sorted(root.rglob('*')):
        if p.is_file(): rows.append({'path':str(p.relative_to(root)),'size':p.stat().st_size,'sha256':sha256_file(p)})
    return rows

def build_suite(root:Path,overwrite:bool=False)->dict[str,Any]:
    if root.exists():
        if not overwrite: raise FileExistsError(root)
        shutil.rmtree(root)
    verified=root/'spreadsheetbench_verified_400'; split_root=root/'spreadsheetbench_id_split'
    verified.mkdir(parents=True)
    records=[]; metadata=[]
    for fam in FAMILIES:
        for pi in range(len(L9_PROFILES)):
            t=build_task(verified,family=fam,profile_index=pi); records.append(t.record); metadata.append(t.metadata)
    records.sort(key=lambda x:x['id']); metadata.sort(key=lambda x:x['id'])
    write_json(verified/'dataset.json',records); write_json(root/'r18_controlled_metadata.json',metadata)
    probes=[]
    for fam in FAMILIES:
        ids=[x['id'] for x in metadata if x['primary_failure_family']==fam]
        probes.extend(select(ids,3,f'{SALT}|{fam}'))
    probes=sorted(probes)
    write_json(root/'r18_split_manifest.json',{'schema_version':'1.0','suite_id':SUITE_ID,'selection_is_outcome_blind':True,'selection_algorithm':'SHA256(salt|family|task_id)','common_heldout_probe':probes,'rules':{'probe_never_fed_to_updater':True,'no_replacement_after_outcome':True}})
    write_json(split_root/'test'/'items.json',[{'id':x} for x in probes])
    rows=file_rows(root)
    manifest={'schema_version':'1.0','suite_id':SUITE_ID,'task_count':54,'selected_probe_count':18,'family_count':6,'families':list(FAMILIES),'block':BLOCK,'role':ROLE,'l9_profiles':[list(x) for x in L9_PROFILES],'files':rows,'dataset_sha256':hashlib.sha256(json.dumps(rows,sort_keys=True,separators=(',',':')).encode()).hexdigest(),'dataset_json_sha256':sha256_file(verified/'dataset.json'),'metadata_sha256':sha256_file(root/'r18_controlled_metadata.json'),'split_manifest_sha256':sha256_file(root/'r18_split_manifest.json')}
    write_json(root/'suite_manifest.json',manifest); return manifest

def self_check(root:Path)->dict[str,Any]:
    meta=json.loads((root/'r18_controlled_metadata.json').read_text())
    split=json.loads((root/'r18_split_manifest.json').read_text())
    assert len(meta)==54 and len(split['common_heldout_probe'])==18 and len(set(split['common_heldout_probe']))==18
    fam={x['id']:x['primary_failure_family'] for x in meta}; counts={f:0 for f in FAMILIES}
    for tid in split['common_heldout_probe']: counts[fam[tid]]+=1
    assert set(counts.values())=={3}
    checked=0
    for row in meta:
        tid=row['id']; base=root/'spreadsheetbench_verified_400'/'spreadsheet'/tid
        wi=load_workbook(base/f'{tid}_init.xlsx',data_only=False); wg=load_workbook(base/f'{tid}_golden.xlsx',data_only=False)
        for q,v in row['golden_answer_cells'].items():
            s,c=q.split('!',1); assert wi[s][c].value is None; assert wg[s][c].value==v
        wi.close(); wg.close(); checked+=1
    return {'status':'PASS_ZERO_PROVIDER','tasks_checked':checked,'selected_probes':18,'per_family':counts,'provider_calls':0}
